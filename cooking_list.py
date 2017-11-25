#-*- coding: UTF-8 -*- 
from openpyxl import load_workbook
import uniout
from itertools import islice
import copy

class register():
	 def __init__(self):
	 	self.name=""
		self.register_type=""
		self.company=""
		self.phone=""
		self.email=""
		self.quantity=""
		self.priority=["","",""]
		self.pay_type=""
		self.special=""
		self.count=0



def get_data():
	# Get data

	for row in islice(ws, 1, None):
		new_register=register()
		new_register.register_type=row[1].value
		print row[1].value
		if row[1].value==u"單人報名":
			i=0
		else:
			i=10

		# Processing various format of name
		new_register.name=row[i+2].value
		new_register.name=new_register.name.replace(u'、',u' ')
		new_register.name=new_register.name.replace(u'/',u' ')
		new_register.name=new_register.name.replace(u'；',u' ')
		new_register.name=new_register.name.replace(u'。',u' ')
		new_register.name=new_register.name.replace(u',',u' ')
		new_register.name=new_register.name.replace(u'，',u' ')
		new_register.name=new_register.name.replace(u'  ',u' ')
		new_register.name=new_register.name.replace(u'   ',u' ')
		new_register.name=new_register.name.replace(u'  ',u' ')
		new_register.name=new_register.name.replace(u'\n',u' ')
		if new_register.name[0]==' ':
			new_register.name=new_register.name[1:]

		#process the count
		# Halu & Austin 被考慮在單人報名裡了。
		if row[1].value==u"單人報名":
			new_register.count=1
		else:
			mycount=new_register.name.count(u' ')+1
			new_register.count=max(2,mycount)


		new_register.company=row[i+3].value
		new_register.phone=row[i+4].value
		new_register.email=row[i+5].value
		new_register.quantity=row[i+6].value
		new_register.priority[0]=row[i+7].value
		new_register.priority[1]=row[i+8].value
		new_register.priority[2]=row[i+9].value
		new_register.pay_type=row[i+10].value
		new_register.special=row[i+11].value

		#print new_register.priority[0]
		x=copy.deepcopy(new_register)
		#print x.priority[0]
		register_list.append(x)
		#print len(register_list)




if __name__ == '__main__':
	wb1=load_workbook("./target.xlsx")
	print wb1.get_sheet_names()
	ws=wb1.active
	register_list=[]
	get_data()
	print len(register_list)
	
	# 6 time slot
	final_list=[[],[],[],[],[],[]]
	for register in register_list:
		print register.name
		print register.count
		#print register.priority[0]

	for register in register_list:
		#print register.priority[0]
		if register.priority[0]==u"12/2(六)早上8:30-12:00(親子班)":
			final_list[0].append(register)
		elif register.priority[0]==u"12/2(六)下午13:00-16:30":
			final_list[1].append(register)
		elif register.priority[0]==u"12/2(六)晚上17:30-21:00":
			final_list[2].append(register)
		elif register.priority[0]==u"12/3(日)早上8:30-12:00":
			final_list[3].append(register)
		elif register.priority[0]==u"12/3(日)下午13:00-16:30":
			final_list[4].append(register)
		elif register.priority[0]==u"12/3(日)晚上17:30-21:00":
			final_list[5].append(register)


	for i in range(6):
		count=0
		for people in final_list[i]:
			count+=people.count
		print count
		name_list=[]
		for people in final_list[i]:
			name_list.append(people.name+u'('+str(people.count)+u')')
		
		s=" "
		s=s.join(name_list)
		print s

	#wb2=Workbook()
	#ws2=wb2.active
	#wb2.save(u"排名單的結果.xlsx")





	


	